"""
Pytest configuration and fixtures for Cortex integration tests.

Provides database setup/teardown, sample data generation, and common assertions.
"""

import os
import pytest
from io import BytesIO
from typing import List, Dict, Any
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import create_engine, event, text
from sqlalchemy.orm import sessionmaker, Session
from sqlalchemy.pool import StaticPool

# Use test database or in-memory SQLite for speed
TEST_DATABASE_URL = os.getenv(
    "TEST_DATABASE_URL",
    "sqlite:///:memory:"
)


@pytest.fixture(scope="function")
def db_engine():
    """Create test database engine."""
    # For SQLite in-memory, use StaticPool to prevent "database is locked" errors
    if "sqlite" in TEST_DATABASE_URL:
        engine = create_engine(
            TEST_DATABASE_URL,
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
    else:
        engine = create_engine(TEST_DATABASE_URL)

    # Create all tables using raw SQL
    # This is simpler than running alembic migrations for tests
    try:
        with engine.begin() as conn:
            # Create weeks table
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS weeks (
                    id INTEGER PRIMARY KEY,
                    season INTEGER NOT NULL,
                    week_number INTEGER NOT NULL CHECK (week_number BETWEEN 1 AND 18),
                    status VARCHAR(20) DEFAULT 'upcoming' CHECK (status IN ('upcoming', 'active', 'completed')),
                    nfl_slate_date DATE,
                    status_override VARCHAR(20),
                    metadata TEXT,
                    is_locked BOOLEAN DEFAULT 0,
                    locked_at TIMESTAMP,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(season, week_number)
                )
            """))

            # Create player_pools table with calibration columns
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS player_pools (
                    id INTEGER PRIMARY KEY,
                    week_id INTEGER NOT NULL REFERENCES weeks(id) ON DELETE CASCADE,
                    player_key VARCHAR(255) NOT NULL,
                    name VARCHAR(255) NOT NULL,
                    team VARCHAR(10) NOT NULL,
                    position VARCHAR(10) NOT NULL CHECK (position IN ('QB', 'RB', 'WR', 'TE', 'DST', 'K')),
                    salary INTEGER NOT NULL CHECK (salary BETWEEN 3000 AND 10000),
                    projection FLOAT CHECK (projection >= 0),
                    ownership FLOAT CHECK (ownership BETWEEN 0 AND 1),
                    ceiling FLOAT,
                    floor FLOAT,
                    source VARCHAR(50) NOT NULL,
                    projection_source VARCHAR(50),
                    opponent_rank_category VARCHAR(20),
                    injury_status VARCHAR(50),
                    projection_floor_original FLOAT,
                    projection_floor_calibrated FLOAT,
                    projection_median_original FLOAT,
                    projection_median_calibrated FLOAT,
                    projection_ceiling_original FLOAT,
                    projection_ceiling_calibrated FLOAT,
                    calibration_applied BOOLEAN DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(week_id, player_key)
                )
            """))

            # Create projection_calibration table
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS projection_calibration (
                    id INTEGER PRIMARY KEY,
                    week_id INTEGER NOT NULL REFERENCES weeks(id) ON DELETE CASCADE,
                    position VARCHAR(10) NOT NULL CHECK (position IN ('QB', 'RB', 'WR', 'TE', 'K', 'DST')),
                    floor_adjustment_percent FLOAT NOT NULL CHECK (floor_adjustment_percent BETWEEN -50 AND 50),
                    median_adjustment_percent FLOAT NOT NULL CHECK (median_adjustment_percent BETWEEN -50 AND 50),
                    ceiling_adjustment_percent FLOAT NOT NULL CHECK (ceiling_adjustment_percent BETWEEN -50 AND 50),
                    is_active BOOLEAN NOT NULL DEFAULT 1,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(week_id, position)
                )
            """))

            # Create weight_profiles table
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS weight_profiles (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name VARCHAR(255) NOT NULL UNIQUE,
                    weights TEXT NOT NULL,
                    config TEXT NOT NULL,
                    is_default BOOLEAN DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """))

            # Create historical_stats table
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS historical_stats (
                    id INTEGER PRIMARY KEY,
                    week_id INTEGER REFERENCES weeks(id) ON DELETE CASCADE,
                    player_name VARCHAR(255) NOT NULL,
                    team VARCHAR(10) NOT NULL,
                    position VARCHAR(10) NOT NULL,
                    week INTEGER CHECK (week BETWEEN 1 AND 18),
                    opponent VARCHAR(10),
                    snaps INTEGER,
                    snap_pct FLOAT,
                    rush_attempts INTEGER,
                    rush_yards INTEGER,
                    rush_tds INTEGER,
                    targets INTEGER,
                    target_share FLOAT,
                    receptions INTEGER,
                    rec_yards INTEGER,
                    rec_tds INTEGER,
                    total_tds INTEGER,
                    touches INTEGER,
                    actual_points FLOAT,
                    salary INTEGER,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """))

            # Create historical_stats_backup table
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS historical_stats_backup (
                    id INTEGER PRIMARY KEY,
                    backup_timestamp TIMESTAMP,
                    data TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """))

            # Create player_aliases table
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS player_aliases (
                    id INTEGER PRIMARY KEY,
                    unmatched_name VARCHAR(255) NOT NULL,
                    canonical_player_key VARCHAR(255) NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(unmatched_name, canonical_player_key)
                )
            """))

            # Create import_history table
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS import_history (
                    id TEXT PRIMARY KEY,
                    week_id INTEGER NOT NULL REFERENCES weeks(id) ON DELETE CASCADE,
                    source VARCHAR(50) NOT NULL,
                    player_count INTEGER NOT NULL,
                    unmatched_count INTEGER DEFAULT 0,
                    metadata TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """))

            # Create player_pool_history table
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS player_pool_history (
                    id INTEGER PRIMARY KEY,
                    import_id TEXT NOT NULL REFERENCES import_history(id) ON DELETE CASCADE,
                    player_key VARCHAR(255) NOT NULL,
                    name VARCHAR(255) NOT NULL,
                    team VARCHAR(10) NOT NULL,
                    position VARCHAR(10) NOT NULL,
                    salary INTEGER NOT NULL,
                    projection FLOAT,
                    ownership FLOAT,
                    ceiling FLOAT,
                    floor FLOAT,
                    source VARCHAR(50) NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """))

            # Create unmatched_players table
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS unmatched_players (
                    id INTEGER PRIMARY KEY,
                    import_id TEXT NOT NULL REFERENCES import_history(id) ON DELETE CASCADE,
                    player_name VARCHAR(255) NOT NULL,
                    team VARCHAR(10),
                    position VARCHAR(10),
                    salary INTEGER,
                    source VARCHAR(50),
                    status VARCHAR(20) DEFAULT 'pending' CHECK (status IN ('pending', 'mapped', 'ignored')),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """))

            # Create week_metadata table (new for week management)
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS week_metadata (
                    id INTEGER PRIMARY KEY,
                    week_id INTEGER NOT NULL UNIQUE REFERENCES weeks(id) ON DELETE CASCADE,
                    season INTEGER NOT NULL,
                    week_number INTEGER NOT NULL CHECK (week_number BETWEEN 1 AND 18),
                    nfl_slate_date DATE NOT NULL,
                    kickoff_time TIME NOT NULL,
                    slate_start_time TIMESTAMP,
                    slate_end_time TIMESTAMP,
                    espn_schedule_url TEXT,
                    import_status VARCHAR(20) DEFAULT 'pending' CHECK (import_status IN ('pending', 'imported', 'error')),
                    import_count INTEGER DEFAULT 0,
                    import_timestamp TIMESTAMP,
                    import_error_message TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(season, week_number)
                )
            """))

            # Create nfl_schedule table (new for week management)
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS nfl_schedule (
                    id INTEGER PRIMARY KEY,
                    season INTEGER NOT NULL,
                    week INTEGER NOT NULL CHECK (week BETWEEN 1 AND 18),
                    slate_date DATE NOT NULL,
                    kickoff_time TIME NOT NULL,
                    game_count INTEGER,
                    is_playoff BOOLEAN DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(season, week)
                )
            """))

            # Create week_status_overrides table (new for week management)
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS week_status_overrides (
                    id INTEGER PRIMARY KEY,
                    week_id INTEGER NOT NULL UNIQUE REFERENCES weeks(id) ON DELETE CASCADE,
                    override_status VARCHAR(20) NOT NULL CHECK (override_status IN ('active', 'upcoming', 'completed')),
                    reason TEXT,
                    overridden_by VARCHAR(255),
                    overridden_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """))

            conn.commit()
    except Exception as e:
        print(f"Error creating tables: {e}")

    yield engine

    # Drop all tables
    try:
        with engine.begin() as conn:
            tables = [
                "week_status_overrides",
                "week_metadata",
                "nfl_schedule",
                "weight_profiles",
                "unmatched_players",
                "player_pool_history",
                "import_history",
                "player_aliases",
                "historical_stats_backup",
                "historical_stats",
                "projection_calibration",
                "player_pools",
                "weeks",
            ]
            for table in tables:
                conn.execute(text(f"DROP TABLE IF EXISTS {table}"))
            conn.commit()
    except Exception as e:
        print(f"Error dropping tables: {e}")

    engine.dispose()


def seed_nfl_schedule(session: Session, year: int = 2025) -> None:
    """Seed NFL schedule data for a year."""
    # NFL schedule data for the season
    schedule_data = []

    # Base dates for 2025 (adjust for other years as needed)
    if year == 2025:
        base_dates = [
            '2025-09-07', '2025-09-14', '2025-09-21', '2025-09-28',
            '2025-10-05', '2025-10-12', '2025-10-19', '2025-10-26',
            '2025-11-02', '2025-11-09', '2025-11-16', '2025-11-23',
            '2025-11-30', '2025-12-07', '2025-12-14', '2025-12-21',
            '2025-12-28', '2026-01-04',
        ]
    else:
        # Generic 18 weeks for other years (Sunday dates)
        from datetime import date, timedelta
        # Find first Sunday of September
        first_sept = date(year, 9, 1)
        first_sunday = first_sept + timedelta(days=(6 - first_sept.weekday()))
        base_dates = [(first_sunday + timedelta(weeks=i)).isoformat() for i in range(18)]

    for week_num in range(1, 19):
        kickoff_time = '12:30' if week_num == 12 else '13:00'  # Thanksgiving

        try:
            session.execute(
                text("""
                    INSERT INTO nfl_schedule (season, week, slate_date, kickoff_time, game_count, is_playoff)
                    VALUES (:season, :week, :slate_date, :kickoff_time, :game_count, :is_playoff)
                """),
                {
                    "season": year,
                    "week": week_num,
                    "slate_date": base_dates[week_num - 1],
                    "kickoff_time": kickoff_time,
                    "game_count": 16,
                    "is_playoff": False,
                }
            )
        except Exception:
            pass  # Skip if already exists

    session.commit()


@pytest.fixture(scope="function")
def db_session(db_engine):
    """Create test database session."""
    connection = db_engine.connect()
    transaction = connection.begin()
    session = sessionmaker(bind=connection, expire_on_commit=False)()

    # Seed NFL schedule for 2025-2027
    for year in [2025, 2026, 2027]:
        seed_nfl_schedule(session, year)

    yield session

    session.close()
    try:
        transaction.rollback()
    except:
        pass
    connection.close()


def create_week(session: Session, season: int = 2024, week_number: int = 1) -> int:
    """Create a week and return its ID."""
    try:
        result = session.execute(
            text("""
                INSERT INTO weeks (season, week_number, status)
                VALUES (:season, :week_number, 'active')
            """),
            {"season": season, "week_number": week_number}
        )
        session.commit()
        return result.lastrowid
    except Exception as e:
        session.rollback()
        # Week might already exist
        result = session.execute(
            text("SELECT id FROM weeks WHERE season = :season AND week_number = :week_number"),
            {"season": season, "week_number": week_number}
        )
        row = result.fetchone()
        return row[0] if row else None


def create_linestar_xlsx() -> BytesIO:
    """Create a sample LineStar XLSX file with 153 players."""
    wb = Workbook()
    ws = wb.active
    ws.title = "LineStar"

    # Headers
    headers = ["Name", "Position", "Team", "Salary", "Projected", "Ceiling", "Floor", "ProjOwn"]
    ws.append(headers)

    # Generate 153 players
    # Simple approach: create enough players to reach 153 total
    positions = ["QB", "RB", "WR", "TE", "DST"]
    teams = ["KC", "LAC", "PHI", "DAL", "CIN", "BAL", "BUF", "NYG", "NO", "TB", "CAR", "GB", "MIN", "SF", "LAR", "TB", "TEN", "LV", "WAS", "DEN", "HOU", "IND"]

    player_count = 0
    for i in range(153):
        team = teams[i % len(teams)]
        position = positions[i % len(positions)]
        name = f"Player{i+1}"
        salary = 5000 + (i % 2000)
        projection = 10.0 + (i % 5)
        ceiling = projection + 2.0
        floor = projection - 2.0
        ownership = 0.05 + (i % 100) / 1000

        ws.append([name, position, team, salary, projection, ceiling, floor, ownership])
        player_count += 1

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_draftkings_xlsx() -> BytesIO:
    """Create a sample DraftKings XLSX file with 174 players."""
    wb = Workbook()
    ws = wb.active
    ws.title = "FE"

    # DraftKings header row (row 1, which we skip)
    ws.append(["Skip Row"])

    # Actual headers (row 2)
    headers = ["Name", "Pos", "T", "S", "Proj", "Ceil", "Flr", "Own", "Notes"]
    ws.append(headers)

    positions = ["QB", "RB", "WR", "TE", "DST"]
    teams = ["KC", "LAC", "PHI", "DAL", "CIN", "BAL", "BUF", "NYG", "NO", "TB", "CAR", "GB", "MIN", "SF", "LAR", "TB", "TEN", "LV", "WAS", "DEN", "HOU", "IND"]

    for i in range(174):
        team = teams[i % len(teams)]
        position = positions[i % len(positions)]
        name = f"Player{i+1}"
        salary = 5000 + (i % 2000)
        projection = 10.0 + (i % 5)
        ceiling = projection + 2.0
        floor = projection - 2.0
        ownership = 0.05 + (i % 100) / 1000
        notes = f"Note{i}"

        ws.append([name, position, team, salary, projection, ceiling, floor, ownership, notes])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_comprehensive_stats_xlsx() -> BytesIO:
    """Create a sample Comprehensive Stats XLSX file with 2690 rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Points"

    headers = [
        "Player", "Tm", "Pos", "Wk", "Opp", "Snaps", "Snp %", "Ratt", "Rsh_yds", "Rsh_td",
        "CTGT", "CTGT%", "Rec", "Rc_yds", "Rc_td", "Tot TD", "Touch", "DK Pts", "Sal"
    ]
    ws.append(headers)

    positions = ["QB", "RB", "WR", "TE", "DST"]
    teams = ["KC", "LAC", "PHI", "DAL", "CIN", "BAL", "BUF", "NYG", "NO", "TB"]

    for i in range(2690):
        week = (i % 18) + 1
        team = teams[i % len(teams)]
        pos = positions[i % len(positions)]
        player_name = f"Stat{i}"
        snaps = 30 + (i % 50)
        snap_pct = (snaps / 80) if snaps > 0 else 0.0
        rush_attempts = i % 20
        rush_yards = rush_attempts * 4
        rush_tds = i % 3
        targets = i % 10
        target_share = (targets / 80) if targets > 0 else 0.0
        receptions = int(targets * 0.65)
        rec_yards = receptions * 8
        rec_tds = i % 2
        total_tds = rush_tds + rec_tds
        touches = rush_attempts + receptions
        actual_points = (rush_yards * 0.1) + (rec_yards * 0.1) + (total_tds * 6)
        salary = 5000 + (i % 2000)

        ws.append([
            player_name, team, pos, week, "OPP", snaps, snap_pct,
            rush_attempts, rush_yards, rush_tds, targets, target_share,
            receptions, rec_yards, rec_tds, total_tds, touches, actual_points, salary
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def verify_import_success(
    session: Session,
    week_id: int,
    expected_count: int,
    source: str
) -> Dict[str, Any]:
    """
    Verify import success by checking database state.

    Returns stats about the import.
    """
    # Count players
    result = session.execute(
        text("SELECT COUNT(*) FROM player_pools WHERE week_id = :week_id AND source = :source"),
        {"week_id": week_id, "source": source}
    )
    count = result.scalar()

    # Check import history
    result = session.execute(
        text("SELECT id, source, player_count FROM import_history WHERE week_id = :week_id AND source = :source ORDER BY created_at DESC LIMIT 1"),
        {"week_id": week_id, "source": source}
    )
    import_record = result.fetchone()

    # Verify player_keys
    result = session.execute(
        text("SELECT COUNT(*) FROM player_pools WHERE week_id = :week_id AND player_key LIKE '%_%_%'"),
        {"week_id": week_id}
    )
    valid_keys = result.scalar()

    # Check ownership normalization
    result = session.execute(
        text("SELECT MIN(ownership), MAX(ownership) FROM player_pools WHERE week_id = :week_id"),
        {"week_id": week_id}
    )
    min_own, max_own = result.fetchone()

    return {
        "player_count": count,
        "expected_count": expected_count,
        "import_record": import_record,
        "valid_player_keys": valid_keys,
        "min_ownership": min_own,
        "max_ownership": max_own,
        "count_matches": count == expected_count,
    }


@pytest.fixture
def linestar_sample() -> BytesIO:
    """Fixture providing LineStar sample file."""
    return create_linestar_xlsx()


@pytest.fixture
def draftkings_sample() -> BytesIO:
    """Fixture providing DraftKings sample file."""
    return create_draftkings_xlsx()


@pytest.fixture
def comprehensive_stats_sample() -> BytesIO:
    """Fixture providing Comprehensive Stats sample file."""
    return create_comprehensive_stats_xlsx()
